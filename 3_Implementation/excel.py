''' Importing Libraries'''
import openpyxl as opx
from openpyxl import Workbook


# pythonProject class with methods
class PythonProject:
    ''' Class '''
    # Reading the Semester marks
    # of Requested PS Number

    data = ''

    @classmethod
    def employee_marks(cls, ps_number):
        ''' Marks '''
        # Input file name
        file_name = 'Employees_Details.xlsx'

        # Excel Sheet Name
        sheet_name = 'Marks'

        try :
            # Reading the input excel file
            excel_obj = opx.load_workbook(file_name)

            # providing the sheet name
            excel_sheet = excel_obj[sheet_name]
        except UnboundLocalError as exception_error:
            print("Not Found", exception_error)
        except FileNotFoundError as exception_error:
            print("Not Found", exception_error)

        # Reading untill max row
        for i in range(1, excel_sheet.max_row + 1):

            # getting current cell value
            cur_cell = str(excel_sheet.cell(row=i, column=1).value)

            # checking is the PS Number is
            # present in the sheet or not
            if ps_number == cur_cell :
                # Getting the current cell column size
                cur_column = int(excel_sheet.cell(row=i, column=1).column)

                # return the next column value
                data = str(excel_sheet.cell(row=i, column=cur_column+1).value)
        return data

    # Reading the Hobbies
    # of Requested PS Number
    @classmethod
    def employee_hobbies(cls, ps_number):
        ''' Hobbies '''
        # Input file name
        file_name = 'Employees_Details.xlsx'

        # Excel Sheet Name
        sheet_name = 'Hobbies'

        try :
            # Reading the input excel file
            excel_obj = opx.load_workbook(file_name)

            # providing the sheet name
            excel_sheet = excel_obj[sheet_name]
        except UnboundLocalError as exception_error:
            print("Not Found", exception_error)
        except FileNotFoundError as exception_error:
            print("Not Found", exception_error)

        # Reading untill max row
        for i in range(1, excel_sheet.max_row + 1):

            # getting current cell value
            cur_cell = str(excel_sheet.cell(row=i, column=1).value)

            # checking is the PS Number is
            # present in the sheet or not
            if ps_number == cur_cell:

                # Getting the current cell column size
                cur_column = int(excel_sheet.cell(row=i, column=1).column)

                # return the next column value
                data = str(excel_sheet.cell(row=i, column=cur_column+1).value)
        return data

    # Reading the Cities Travelled
    # by Requested PS Number
    @classmethod
    def employee_cities_travelled(cls, ps_number):
        ''' Cities Travelled '''
        # Input file name
        file_name = 'Employees_Details.xlsx'

        # Excel Sheet Name
        sheet_name = 'Cities Travelled'

        try :
            # Reading the input excel file
            excel_obj = opx.load_workbook(file_name)

            # providing the sheet name
            excel_sheet = excel_obj[sheet_name]
        except UnboundLocalError as exception_error:
            print("Not Found", exception_error)
        except FileNotFoundError as exception_error:
            print("Not Found", exception_error)

        # Reading untill max row
        for i in range(1, excel_sheet.max_row + 1):

            # getting current cell value
            cur_cell = str(excel_sheet.cell(row=i, column=1).value)

            # checking is the PS Number is
            # present in the sheet or not
            if ps_number == cur_cell:

                # Getting the current cell column size
                cur_column = int(excel_sheet.cell(row=i, column=1).column)

                # return the next column value
                data = str(excel_sheet.cell(row=i, column=cur_column+1).value)
        return data

    # Reading the Programming Languages
    # of Requested PS Number
    @classmethod
    def employee_programming(cls, ps_number):
        ''' Programming Language '''
        # Input file name
        file_name = 'Employees_Details.xlsx'

        # Excel Sheet Name
        sheet_name = 'Programming Language'

        try :
            # Reading the input excel file
            excel_obj = opx.load_workbook(file_name)

            # providing the sheet name
            excel_sheet = excel_obj[sheet_name]
        except UnboundLocalError as exception_error:
            print("Not Found", exception_error)
        except FileNotFoundError as exception_error:
            print("Not Found", exception_error)

        # Reading untill max row
        for i in range(1, excel_sheet.max_row + 1):

            # getting current cell value
            cur_cell = str(excel_sheet.cell(row=i, column=1).value)

            # checking is the PS Number is
            # present in the sheet or not
            if ps_number == cur_cell:

                # Getting the current cell column size
                cur_column = int(excel_sheet.cell(row=i, column=1).column)

                # return the next column value
                data = str(excel_sheet.cell(row=i, column=cur_column+1).value)
        return data

    # Reading the Domain
    # of Requested PS Number
    @classmethod
    def employee_domain(cls, ps_number):
        ''' Domain '''
        # Input file name
        file_name = 'Employees_Details.xlsx'

        # Excel Sheet Name
        sheet_name = 'Domain'

        try :
            # Reading the input excel file
            excel_obj = opx.load_workbook(file_name)

            # providing the sheet name
            excel_sheet = excel_obj[sheet_name]
        except UnboundLocalError as exception_error:
            print("Not Found", exception_error)
        except FileNotFoundError as exception_error:
            print("Not Found", exception_error)

        # Reading untill max row
        for i in range(1, excel_sheet.max_row + 1):

            # getting current cell value
            cur_cell = str(excel_sheet.cell(row=i, column=1).value)

            # checking is the PS Number is
            # present in the sheet or not
            if ps_number == cur_cell:

                # Getting the current cell column size
                cur_column = int(excel_sheet.cell(row=i, column=1).column)

                # return the next column value
                data = str(excel_sheet.cell(row=i, column=cur_column+1).value)
        return data

    # writing the requested data
    # to a new excel file
    # pylint: disable=too-many-arguments
    @classmethod
    def employee_new_excel(
            cls,
            ps_number,
            employee_marks,
            employee_hobbies,
            employee_cities,
            employee_programming,
            employee_domain
    ):
        ''' New File '''
        # creating a openpyxl workbook instance
        work_book = Workbook()

        # output file
        dest_file_name = ps_number + '-Data.xlsx'

        # getting active sheet
        ws1 = work_book.active

        # saving the PS number as the sheet name
        ws1.title = ps_number

        # adjusting the width of A to 15
        ws1.column_dimensions['A'].width = 15

        # adding PS number string to A1 cell
        ws1['A1'] = 'PS Number'

        # adding PS number to A2 cell
        ws1['A2'] = ps_number

        # adjusting the width of B to 15
        ws1.column_dimensions['B'].width = 15

        # adding Marks string to B1 cell
        ws1['B1'] = 'Marks'

        # adding Marks string to B2 cell
        ws1['B2'] = employee_marks

        # adjusting the width of C to 40
        ws1.column_dimensions['C'].width = 40

        # adding Hibbies string to C1 cell
        ws1['C1'] = 'Hobbies'

        # adding Hobbies to C2 cell
        ws1['C2'] = employee_hobbies

        # adjusting the width of D to 30
        ws1.column_dimensions['D'].width = 30

        # adding Cities Travelled string to D1 cell
        ws1['D1'] = 'Cities Travelled'

        # adding Cities Travelled to D2 cell
        ws1['D2'] = employee_cities

        # adjusting the width of E to 40
        ws1.column_dimensions['E'].width = 40

        # adding Programming Language string to E1 cell
        ws1['E1'] = 'Programming Language'

        # adding Programming Language to E2 cell
        ws1['E2'] = employee_programming

        # adjusting the width of F to 35
        ws1.column_dimensions['F'].width = 35

        # adding Domain string to F1 cell
        ws1['F1'] = 'Domain'

        # adding Domain to F2 cell
        ws1['F2'] = employee_domain

        try:
            # saving the new excel file
            # with requested data
            work_book.save(filename = dest_file_name)
        except UnboundLocalError as exception_error:
            print(exception_error)
        except FileNotFoundError as exception_error:
            print("Not Found", exception_error)

        return 0
